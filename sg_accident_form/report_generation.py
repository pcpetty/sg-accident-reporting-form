# SG Accident Report Generation

# Import Libraries and Modules
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import json
from openpyxl.chart import BarChart, Reference
import datetime
from db_operations import fetch_driver_name, fetch_vehicle_plate
from fpdf import FPDF

def export_to_excel(data, filename="Accident_Report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Accident Report"

    # Add headers
    headers = ["Reference Key", "Accident Date", "Accident Time", "Location", "Hazmat", "Driver Name", "Vehicle Plate"]
    ws.append(headers)

    # Fetch driver name and vehicle plate based on IDs
    driver_name = fetch_driver_name(data.get("driver_id")) if data.get("driver_id") else "Unknown"
    vehicle_plate = fetch_vehicle_plate(data.get("vehicle_id")) if data.get("vehicle_id") else "Unknown"

    # Add data
    ws.append([
        data["reference_key"],
        data.get("accident_date", "Unknown"),  # Default to "Unknown" if key doesn't exist
        data.get("accident_time", "Unknown"),
        data.get("accident_location", "Unknown"),
        "Yes" if data.get("hazmat", False) else "No",
        driver_name,
        vehicle_plate,
    ])


    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Add chart (example: Hazmat incidents)
    chart = BarChart()
    chart.title = "Hazmat Incidents"
    chart.x_axis.title = "Reference Key"
    chart.y_axis.title = "Count"
    data = Reference(ws, min_col=5, min_row=1, max_row=ws.max_row, max_col=5)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "H10")

    wb.save(filename)
    print(f"Excel report saved as {filename}.")

# ---- EXPORT TO PDF ----#

from fpdf import FPDF

def export_to_pdf(data, filename="Accident_Report.pdf"):
    """
    Exports the accident report to a structured PDF format.
    """
    pdf = FPDF()
    pdf.add_page()
    # Title
    pdf.set_font("Arial", style="B", size=16)
    pdf.cell(200, 10, txt="Accident Report", ln=True, align="C")
    pdf.ln(10)  # Add a line break

    # Helper function to format sections
    def add_section(title, content):
        pdf.set_font("Arial", style="B", size=14)
        pdf.cell(200, 10, txt=title, ln=True)
        pdf.set_font("Arial", size=12)
        for key, value in content.items():
            pdf.cell(200, 10, txt=f"{key}: {value}", ln=True)
        pdf.ln(5)  # Add a small line break after the section

    # Section: Report Details
    report_details = {
        "Reference Key": data.get("reference_key", "N/A"),
        "Accident Date": data.get("accident_date", "N/A"),
        "Accident Time": data.get("accident_time", "N/A"),
        "Location": data.get("accident_location", "N/A"),
        "Hazmat Involved": "Yes" if data.get("hazmat", False) else "No",
    }
    add_section("Report Details", report_details)

    # Section: V1 Driver
    v1_driver = data.get("v1_driver") or {}
    v1_driver_details = {
        "Name": v1_driver.get("driver_name", "N/A"),
        "Phone": v1_driver.get("driver_phone", "N/A"),
        "License Number": v1_driver.get("license_number", "N/A"),
        "License Expiry": v1_driver.get("license_expiry", "N/A"),
        "Injured": "Yes" if v1_driver.get("driver_injury", False) else "No",
    }
    add_section("V1 Driver Details", v1_driver_details)

    # Section: V1 Vehicle
    v1_vehicle = data.get("v1_vehicle") or {}  # Safeguard against None
    v1_vehicle_details = {
        "License Plate": v1_vehicle.get("license_plate", "N/A"),
        "Make": v1_vehicle.get("make", "N/A"),
        "Model": v1_vehicle.get("model", "N/A"),
        "Year": v1_vehicle.get("year", "N/A"),
        "Color": v1_vehicle.get("color", "N/A"),
    }
    add_section("V1 Vehicle Details", v1_vehicle_details)

    # Section: V2 Driver
    v2_driver = data.get("v2_driver") or {}
    v2_driver_details = {
        "Name": v2_driver.get("driver_name", "N/A"),
        "Phone": v2_driver.get("driver_phone", "N/A"),
        "License Number": v2_driver.get("license_number", "N/A"),
        "License Expiry": v2_driver.get("license_expiry", "N/A"),
        "Injured": "Yes" if v2_driver.get("driver_injury", False) else "No",
    }
    add_section("V2 Driver Details", v2_driver_details)

    # Section: V2 Vehicle
    v2_vehicle = data.get("v2_vehicle") or {}
    v2_vehicle_details = {
        "License Plate": v2_vehicle.get("license_plate", "N/A"),
        "Make": v2_vehicle.get("make", "N/A"),
        "Model": v2_vehicle.get("model", "N/A"),
        "Year": v2_vehicle.get("year", "N/A"),
        "Color": v2_vehicle.get("color", "N/A"),
    }
    add_section("V2 Vehicle Details", v2_vehicle_details)

    # Section: Additional Remarks
    additional_remarks = data.get("additional_remarks", "No additional remarks provided.")
    add_section("Additional Remarks", {"Remarks": additional_remarks})

    # Save the PDF
    pdf.output(filename)
    print(f"PDF report saved as {filename}.")

# ---- ---- # 

def save_to_json(data, filename="accident_report.json"):
    def custom_serializer(obj):
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return obj.isoformat()  # Convert date/datetime to ISO 8601 string
        elif isinstance(obj, set):
            return list(obj)  # Convert sets to lists
        raise TypeError(f"Type {type(obj)} not serializable")

    try:
        with open(filename, "w") as file:
            json.dump(data, file, indent=4, default=custom_serializer)
        print(f"Data successfully saved to {filename}")
    except Exception as e:
        print(f"Error saving to JSON: {e}")
